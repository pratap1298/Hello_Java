# -*- coding: utf-8 -*-
"""
Created on Thu Oct 22 17:28:41 2020

@author: 320106402
"""

# -*- coding: utf-8 -*-
"""
Created on Sat Oct 17 13:00:46 2020

@author: 320106402
"""

import openpyxl  as op
import csv 
import pandas as pd 


'''
path = "C:\\Users\\320106402\\Downloads\\CEA\\imcs-billing-report-2020-09-14.xlsx"

path1 = "C:\\Users\\320106402\\Downloads\\CEA\\HSDP-Trend.xlsx"
'''

#C:\\Users\\320106402\\Downloads\\CEA\\imcs-billing-report-2020-09-14.xlsx
#C:\\Users\\320106402\\Downloads\\CEA\\output.xlsx
#C:\\Users\\320106402\\Downloads\\CEA\\HSDP-Trend.xlsx

user_input = input("Enter imcs-billing-report file path : ")
user_input1=input("Enter HSDF Trend file path : ")


#print(user_input[-15:-7])
a=user_input[-15:-7]
b="01"
date=a+b
print(date)

input1=op.load_workbook(user_input)
cloud_apps=input1['Cloud Foundry App Run Time']
row_count = cloud_apps.max_row
print("Max row in cloud_apps -> row_count")
print(row_count)
output=op.load_workbook(user_input1)
sheet = output['HSDP Bill']
row_count1 = sheet.max_row
print("Max row before insterting  cloud_apps -> row_count1")
print(row_count1)
a=(cloud_apps.cell(row=2,column=5).value)*1.15
#print(a)
b=cloud_apps.cell(row=2,column=3).value
#print(b)

 
'''
-----------------------------------------------------------------------------
code for cloud foundry app run time
-----------------------------------------------------------------------------
'''

for i in range (1,row_count+1):  
    if cloud_apps.cell(row=i,column=3).value=='ASP-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 , 2)
    elif cloud_apps.cell(row=i,column=3).value=='BU-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Integration'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='CAT-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='CPP-DECOMMISSION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)     
    elif cloud_apps.cell(row=i,column=3).value=='IOTCLOUD-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox' 
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='IOTCLOUD-QA':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='IoT-SystemTest':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='MTS-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='MTS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='OPS_DevOps':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='PF-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox' 
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='production-bap':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='production-cat':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='production-RDW':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='PRS2.0-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox' 
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='PRS2.0-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Integration' 
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RCW-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSM-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSM-QA':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='RSM'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSW-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSW-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging' 
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='TEST-AUTOMATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='UAT':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='PF-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSW-STAGING-V2':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='RSW-STAGING-V3':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='PRS-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='ASP-Redshift-Retention':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='ASP-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    elif cloud_apps.cell(row=i,column=3).value=='DAW-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='Cloud Foundry'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Integration'
        sheet.cell(row=1+row_count1,column=6).value='APPS'
        sheet.cell(row=1+row_count1,column=5).value=cloud_apps.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_apps.cell(row=i,column=5).value)*1.15 ,2)
    

        



'''
-----------------------------------------------------------------------------
code for cloud foundry service run time -->DWH
-----------------------------------------------------------------------------
'''



DWH=input1['Cloud Foundry Service Run Time']
row_count = DWH.max_row
#print("Max row in DHW -> row_count")
#print(row_count)
row_count1 = sheet.max_row
#print("Max row before insterting  DHW -> row_count1")
#print(row_count1)
for i in range (1,row_count+1):
    if DWH.cell(row=i,column=3).value=='ASP-STAGING' and DWH.cell(row=i,column=4).value=='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='DWH'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=5).value='ASP-STAGING'
        sheet.cell(row=1+row_count1,column=6).value='Redshift'
        sheet.cell(row=1+row_count1,column=7).value=round((DWH.cell(row=i,column=8).value)*1.15 ,2)
    elif DWH.cell(row=i,column=3).value=='CPP-DECOMMISSION' and DWH.cell(row=i,column=4).value=='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='DWH'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=5).value='CPP-DECOMMISSION'
        sheet.cell(row=1+row_count1,column=6).value='Redshift'
        sheet.cell(row=1+row_count1,column=7).value=round((DWH.cell(row=i,column=8).value)*1.15 ,2)
    elif DWH.cell(row=i,column=3).value=='production-bap' and DWH.cell(row=i,column=4).value=='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='DWH'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=5).value='production-bap'
        sheet.cell(row=1+row_count1,column=6).value='Redshift'
        sheet.cell(row=1+row_count1,column=7).value=round((DWH.cell(row=i,column=8).value)*1.15 ,2)
    elif DWH.cell(row=i,column=3).value=='ASP-MR-RETENTION' and DWH.cell(row=i,column=4).value=='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='DWH'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=5).value='ASP-MR-RETENTION'
        sheet.cell(row=1+row_count1,column=6).value='Redshift'
        sheet.cell(row=1+row_count1,column=7).value=round((DWH.cell(row=i,column=8).value)*1.15 ,2)
    elif DWH.cell(row=i,column=3).value=='ASP-Redshift-Retention' and DWH.cell(row=i,column=4).value=='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='DWH'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=5).value='ASP-MR-RETENTION'
        sheet.cell(row=1+row_count1,column=6).value='Redshift'
        sheet.cell(row=1+row_count1,column=7).value=round((DWH.cell(row=i,column=8).value)*1.15 ,2)



'''

-----------------------------------------------------------------------------
code for EC2
-----------------------------------------------------------------------------

'''


EC2=input1['Elastic Compute Cloud']
row_count = EC2.max_row
#print("Max row in EC2 -> row_count")
#print(row_count)
row_count1 = sheet.max_row
#print("Max row before insterting  EC2 -> row_count1")
#print(row_count1)

for i in range (1,row_count+1):
    if EC2.cell(row=i,column=4).value=='asp-313-worker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-314-plotter01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-logdrain01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker02.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker03.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker04.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker05.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker06.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker07.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker08.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker09.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-production-worker10.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='hdp-dsp-chdocker.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='hdp-imcs-dawpoc.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='hdp-imcs-devops.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='hdp-imcs-iot-docker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='imcs-dsp-dev01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='ps-asp-chpoc1.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='whs-imcsweb01':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='whs-imcsweb02':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='whs-imcsweb03':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='whs-imcsweb04':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='wrp-imcsweb01':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='wrp-imcsweb02':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='wrp-imcsweb03':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-vertica-mc.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-vertica-worker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-vertica-worker02.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-vertica-worker03.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-docker-worker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-nomad-worker01.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-nomad-worker02.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-nomad-worker02.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-nomad-worker03.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-edi-poc.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='hdp-imcs-performancebridge.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    elif EC2.cell(row=i,column=4).value=='asp-docker-worker02.dev':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='EC2'
        sheet.cell(row=1+row_count1,column=5).value=EC2.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value=EC2.cell(row=i,column=5).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((EC2.cell(row=i,column=8).value)*1.15,2)
    
    
    




'''

-----------------------------------------------------------------------------
code for cloud foundry service run time
-----------------------------------------------------------------------------

'''

cloud_serv=input1['Cloud Foundry Service Run Time']
row_count = cloud_serv.max_row
#print("Max row in cloud service -> row_count")
#print(row_count)
row_count1 = sheet.max_row
#print("Max row before insterting  cloud service -> row_count1")
#print(row_count1)

for i in range (1,row_count+1):
    if cloud_serv.cell(row=i,column=3).value=='ASP-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='ASP-STAGING' and cloud_serv.cell(row=i,column=4).value!='Redshift' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='BU-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Integration'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='CAT-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='CPP-DECOMMISSION' and cloud_serv.cell(row=i,column=4).value!='Redshift':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='IOTCLOUD-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='IOTCLOUD-QA':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='IoT-SystemTest':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='MTS-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='MTS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='OPS_DevOps':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='PF-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='production-bap' and cloud_serv.cell(row=i,column=4).value!='Redshift':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='production-cat':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='production-RDW':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='PRS2.0-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='PRS2.0-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='RSM-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='RSM'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='RSM-QA':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='RSM'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='RSW-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='RSW-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='TEST-AUTOMATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='UAT':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='DataScience'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='PF-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)        
    elif cloud_serv.cell(row=i,column=3).value=='RSW-STAGING-V2':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='PRS-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='RSW-STAGING-V3':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='ASP-INTEGRATION':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
    elif cloud_serv.cell(row=i,column=3).value=='DAW-STAGING':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='ServiceBroker'
        sheet.cell(row=1+row_count1,column=5).value=cloud_serv.cell(row=i,column=3).value
        sheet.cell(row=1+row_count1,column=6).value=cloud_serv.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Integration'
        sheet.cell(row=1+row_count1,column=7).value=round((cloud_serv.cell(row=i,column=8).value)*1.15,2)
     


'''

-----------------------------------------------------------------------------
code for cloud foundry S3
-----------------------------------------------------------------------------

'''


S3=input1['S3']
row_count = S3.max_row
#print("Max row in S3-> row_count")
#print(row_count)
row_count1 = sheet.max_row
#print("Max row before insterting  S3 -> row_count1")
#print(row_count1)


for i in range (1,row_count+1):
    if S3.cell(row=i,column=4).value=='CAT-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='CPP-DECOMMISSION' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='CT-RADAR'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='IOTCLOUD-DEV'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='OPS_DevOps'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='production-bap'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='production-cat'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='production-RDW'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='UAT'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='DataScience'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    elif S3.cell(row=i,column=4).value=='RSW-STAGING-V2'  :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='S3'
        sheet.cell(row=1+row_count1,column=5).value=S3.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Remote'
        sheet.cell(row=1+row_count1,column=3).value='Staging'
        sheet.cell(row=1+row_count1,column=4).value='DataScience'
        sheet.cell(row=1+row_count1,column=7).value=round((S3.cell(row=i,column=8).value)*1.15 ,2)
    
  

'''

-----------------------------------------------------------------------------
RDS
-----------------------------------------------------------------------------

'''

RDS=input1['RDS']
row_count = RDS.max_row
#print("Max row in RDS-> row_count")
#print(row_count)
row_count1 = sheet.max_row
#print("Max row before insterting  RDS -> row_count1")
#print(row_count1)


for i in range (1,row_count+1):
    if RDS.cell(row=i,column=4).value=='ASP-DEV':
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='ASP-STAGING' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='BU-INTEGRATION' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Integration'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='CAT-DEV' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='CPP-DECOMMISSION' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='IOTCLOUD-DEV' :
        row_count1 = sheet.max_row 
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='IOTCLOUD-QA' :
        row_count1 = sheet.max_row 	
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='IOT'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='IoT-SystemTest' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='MTS-DEV' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='MTS'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='OPS_DevOps' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Ops'
        sheet.cell(row=1+row_count1,column=4).value='NA'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='PF-DEV' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='PF-STAGING' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='PF'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='production-bap' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='ASP'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='production-cat' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='CAT/DAW'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='production-RDW' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Production'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='PRS-STAGING' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='PRS'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='RSM-DEV' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='RSM'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='RSM-QA' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='RSM'
        sheet.cell(row=1+row_count1,column=4).value='QA'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='RSW-DEV' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='RSW-STAGING-V3' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Staging'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='TEST-AUTOMATION' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='Sandbox'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)
    elif RDS.cell(row=i,column=4).value=='UAT' :
        row_count1 = sheet.max_row
        sheet.cell(row=1+row_count1,column=1).value=date
        sheet.cell(row=1+row_count1,column=2).value='RDS'
        sheet.cell(row=1+row_count1,column=5).value=RDS.cell(row=i,column=4).value
        sheet.cell(row=1+row_count1,column=6).value='Storage'
        sheet.cell(row=1+row_count1,column=3).value='Remote'
        sheet.cell(row=1+row_count1,column=4).value='DataScience'
        sheet.cell(row=1+row_count1,column=7).value=round((RDS.cell(row=i,column=8).value)*1.15 ,2)


 
output.save(user_input1)

'''
output.save('C:\\Users\\320106402\\Downloads\\CEA\\HSDP-Trend.xlsx')
'''

'''

with open('finaloutput.csv', 'w', newline="") as outfile:
    col = csv.writer(outfile) 
    for r in sheet.rows: 
        col.writerow([cell.value for cell in r]) 
    outfile.close()
    
    
df= pd.read_csv(outfile)
df.to_csv('finaloutput5.csv',index=False)
'''
'''
df1=pd.read_excel(user_input1,'HSDP Bill')
df1["Environment"].fillna("NA", inplace = True) 
aa=df1.loc[17,['Month']]
print(df1)
print(df1.dtypes)
print(aa)
df1.to_csv('finaloutput.csv',index=False)
'''

df=pd.read_excel(user_input1,'HSDP Bill')
df["Environment"].fillna("NA", inplace = True) 
#df['Month'] = pd.to_datetime(df['Month'], format='%m/%d/%Y %H:%M:%S').dt.strftime('%Y-%m-%d')
#df['Month'] = df['Month'].astype('datetime64[ns]')
df['Month'] = pd.to_datetime(df['Month'])
df['Month'] = df['Month'].dt.strftime('%Y-%m-%d')
df.to_csv('HSDP-Trend.csv',index=False)
df.to_excel('finaloutput.xlsx',index=False)







print("final output .csv file is saved in the HSDP Trend file location path ")
  
#C:\\Users\\320106402\\Downloads\\CEA\\imcs-billing-report-2020-09-14.xlsx
#C:\\Users\\320106402\\Downloads\\CEA\\output.xlsx
#C:\\Users\\320106402\\Downloads\\CEA\\HSDP-Trend.xlsx





   


